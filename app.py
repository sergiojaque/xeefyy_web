import hashlib
import sys

import xmljson as xmljson
import xmltodict
from pivot import pivot
from reportlab.lib.styles import getSampleStyleSheet
from bs4 import BeautifulSoup
from flask import Flask, render_template, request, redirect, flash, session, url_for, Response
from flask_sqlalchemy import SQLAlchemy
import psycopg2
from psycopg2 import connect
import psycopg2.extensions as _ext
from sqlalchemy.sql.functions import current_user
from werkzeug.security import check_password_hash, generate_password_hash
import os
import zipfile
import xlrd
import tempfile
import openpyxl
import string
import urllib3
import pdfkit
from datetime import datetime, timedelta
from xbrl import XBRLParser, GAAP, GAAPSerializer

dbdir = "sqlite:///" + "xeeffy.db"

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = dbdir
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db_user = os.environ.get('DBAAS_USER_NAME', 'jcarrasco')
db_password = os.environ.get('DBAAS_USER_PASSWORD', '1234')
db_connect = os.environ.get('DBAAS_DEFAULT_CONNECT_DESCRIPTOR', "190.114.255.158")
service_port = port = os.environ.get('PORT', '5432')
cnx = connect(
    database='xeeffy',
    host='190.114.255.158',
    port='5432',
    user='jcarrasco',
    password='1234')

status = cnx.get_transaction_status()
if status == _ext.TRANSACTION_STATUS_UNKNOWN:
    # server connection lost
    cnx.close()
elif status != _ext.TRANSACTION_STATUS_IDLE:
    # connection in error or in transaction
    cnx.rollback()
db = SQLAlchemy(app)


class PoolError(psycopg2.Error):
    pass


class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    usuario = db.Column(db.String(50), unique=True)
    contrasena = db.Column(db.String(50))


@app.route('/login', methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if cnx.status == 1:
            username = request.form["username"]
            password = request.form["password"]
            user = Usuario.query.filter_by(usuario=username).first()
            if user and check_password_hash(user.contrasena, request.form["password"]):
                success_message = 'Bienvenido {}'.format(username)
                flash(success_message)
                session['username'] = username
                return redirect("/index")
            else:
                error_message = 'usuario o contraseña no valido'
                flash(error_message)
        else:
            flash("error al conectar bd")
            return redirect("/login")
    return render_template("/login.html")


@app.route("/logout", methods=["GET", "POST"])
def logout():
    session.clear()
    session.pop("username", None)
    session["username"] = "unknown"
    return redirect("/login")


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        hashed_pw = generate_password_hash(request.form["password"], method="sha256")
        new_user = Usuario(usuario=request.form["username"], contrasena=hashed_pw)
        db.session.add(new_user)
        db.session.commit()
        return redirect("/login")
    return render_template("signup.html")


app.secret_key = "9;~X!cp4KhL9u}4#"


@app.route('/', methods=["GET", "POST"])
@app.route('/index', methods=["GET", "POST"])
def index():
    return render_template("/index.html")


@app.route('/contacts')
def contacts():
    return render_template("/contacts.html")


@app.route('/analisis')
def analisis():
    return render_template("/analisis.html")


@app.before_request
def session_management():
    session.permanent = True


@app.route('/entidades', methods=["GET", "POST"])
def entidades():
    if session["username"] != 'unknown':
        cur = cnx.cursor()
        cur.execute("select * from entidad")
        data = cur.fetchall()
        return render_template("/entidades.html", dato=data)
    return redirect("login")


@app.route('/roles', methods=["GET", "POST"])
def roles():
    if session["username"] != 'unknown':
        cur = cnx.cursor()
        cur.execute("select * from role")
        data = cur.fetchall()
        return render_template("/roles.html", dato=data)
    return redirect("login")


@app.route('/roleCuenta/<cod_role>', methods=["GET", "POST"])
def rolesCuenta(cod_role):
    if session["username"] != 'unknown':
        cur = cnx.cursor()
        cur.execute("select * from cuenta join role_cuenta using (cod_cuenta) where cod_role ='%s'" % cod_role)
        data = cur.fetchall()
        return render_template("/roleCuenta.html", dato=data)
    return redirect("login")


@app.route('/reportes/<cod_documento>/<cod_role>', methods=["GET", "POST"])
def reportes(cod_documento, cod_role):
    if session["username"] != 'unknown':
        sql = cnx.cursor()
        sql.execute("""select desc_role, cod_role_cuenta,desc_escenario, desc_dimension, desc_tipo_cuenta,coalesce(desc_role_cuenta, etiqueta, desc_cuenta) desc_cuenta, cod_periodo,desc_periodo, desc_detalle_documento, decimales, cod_unidad, desc_unidad, cod_entidad,desc_entidad, case when not decimales isnull then desc_detalle_documento::numeric / 10 ^ abs(decimales) else 0 end valor 
                    from detalle_documento 
                    join cuenta using (cod_cuenta) 
                    left join role_cuenta using (cod_cuenta)
                    left join role using(cod_role)
                    left join unidad using (cod_unidad)
                    join tipo_cuenta using (cod_tipo_cuenta)
                    join contexto using (cod_contexto) 
                    join entidad using (cod_entidad)
                    join periodo using (cod_periodo) 
                    left join context_escenarios using (cod_contexto) 
                    left join escenario using (cod_escenario) 
                    left join dimension using (cod_dimension) 
                    --where cod_periodo = 44 and cod_tipo_cuenta in (6,7)
                    where cod_dimension is null  and cod_documento = %s and cod_role =%s
                    order by cod_role_cuenta, desc_cuenta""" % (cod_documento, cod_role))
        data = sql.fetchall()
        if data[0] == []:
            flash("No Existe")
            return render_template("/detalle_documento")
        for d in data:
            desc_role = d[0]
            entidad = d[13]
            cuenta = d[5]
            dineros = d[8]
            fechas = d[7]

        return render_template("/reportes.html", data=data, desc_role=desc_role, entidad=entidad, cuenta=cuenta,
                               dineros=dineros, fechas=fechas)
    return redirect("login")


@app.route('/documentos', methods=["GET", "POST"])
def documentos():
    if session["username"] != 'unknown':
        cur = cnx.cursor()
        cur.execute("""SELECT
        public.entidad.cod_entidad,
        public.entidad.desc_entidad,
        public.documento.desc_documento,
        public.documento.anio,
        public.documento.mes,
        public.documento.cod_documento,
        public.documento.hash
        FROM
        public.documento
        INNER JOIN public.entidad ON public.documento.cod_entidad = public.entidad.cod_entidad""")
        data = cur.fetchall()
        return render_template("/documentos.html", dato=data)
    return redirect("login")


@app.route('/detalle_documento/<cod_documento>', methods=["GET", "POST"])
def detalle_documentos(cod_documento):
    if session["username"] != 'unknown':
        cur1 = cnx.cursor()
        cur = cnx.cursor()
        cur1.execute("""SELECT
public.role.cod_role,
public.role.desc_role
FROM
public.role""")
        cur.execute("""SELECT
public.detalle_documento.cod_documento,
public.cuenta.desc_cuenta
FROM
public.cuenta
INNER JOIN public.detalle_documento ON public.detalle_documento.cod_cuenta = public.cuenta.cod_cuenta
WHERE
public.detalle_documento.cod_documento ='%s'""" % cod_documento)
        data = cur.fetchall()
        data1 = cur1.fetchall()
        if data == []:
            flash("NO ENCONTRADO")
            return redirect("/documentos")
        else:
            return render_template("/detalle_documento.html", dato1=data1, dato=data, cod_documento=cod_documento)
    return redirect("login")


# @app.route('/ndocumento', methods=["GET", "POST"])
# def ndocumento():
#     if session["username"] != 'unknown':
#         if request.method == "POST":
#             entidad = request.args.get("entidad")
#             desc_documento = request.args.get("desc_documento")
#             desc_detalle_documento = request.args.get("desc_detalle_documento")
#             ano = request.args.get("ano")
#             mes = request.args.get("mes")
#             hash = request.args.get("hash")
#             cuenta = request.args.get("cuenta")
#             unidad = request.args.get("unidad")
#             decimal = request.args.get("decimal")
#             valor = request.args.get("valor")
#             contexto = request.args.get("contexto")
#             cur = cnx.cursor()
#             cur1 = cnx.cursor()
#             cur2 = cnx.cursor()
#             cur3 = cnx.cursor()
#             cur4 = cnx.cursor()
#             cur5 = cnx.cursor()
#             cur.execute("""INSERT INTO documento(desc_documento,anio,mes,hash)
#                             VALUES('%s','%s','%s','%s')""" % (desc_documento, ano, mes, hash))
#             dato = cur.fetchall()
#             cur1 = cnx.execute("""INSERT INTO entidad(desc_entidad)
#                 VALUES ('%s')""" % (entidad))
#             dato1 = cur1.fetchall()
#             cur2.execute(
#                 """INSERT INTO detalle_documento(desc_detalle_documento,decimales,valor) VALUES ('%s','%s','%s')""" % (
#                 desc_detalle_documento, decimal, valor))
#             dato2 = cur2.fetchall()
#             cur3.execute("""INSERT INTO unidad(desc_unidad) VALUES ('%s')""" % unidad)
#             dato3 = cur3.fetchall()
#             cur4.execute("""INSERT INTO contexto(desc_contexto) VALUES ('%s')""" % contexto)
#             dato4 = cur4.fetchall()
#             cur5.execute("""INSERT INTO cuenta(desc_cuenta) VALUES ('%s','%s')""" % cuenta)
#             dato5 = cur5.fetchall()
#             dato.commit()
#             dato1.commit()
#             dato2.commit()
#             dato3.commit()
#             dato4.commit()
#             dato5.commit()
#             return render_template("/ndocumento.html",dato1=dato1, dato2=dato2, dato3=dato3, dato4=dato4,
#                                    dato5=dato5)
#         cur = cnx.cursor()
#         cur.execute("SELECT desc_documento from documento")
#         dato = cur.fetchall()
#         return render_template("/ndocumento.html",dato=dato)
#     return redirect("/login")

@app.route('/pdf/<cod_documento>/<cod_role>', methods=["GET", "POST"])
def pdf(cod_documento, cod_role):
    cursor = cnx.cursor()
    cursor.execute("""select desc_role, cod_role_cuenta,desc_escenario, desc_dimension, desc_tipo_cuenta,coalesce(desc_role_cuenta, etiqueta, desc_cuenta) desc_cuenta, cod_periodo,desc_periodo, desc_detalle_documento, decimales, cod_unidad, desc_unidad, cod_entidad,desc_entidad, case when not decimales isnull then desc_detalle_documento::numeric / 10 ^ abs(decimales) else 0 end valor 
                            from detalle_documento 
                            join cuenta using (cod_cuenta) 
                            left join role_cuenta using (cod_cuenta)
                            left join role using(cod_role)
                            left join unidad using (cod_unidad)
                            join tipo_cuenta using (cod_tipo_cuenta)
                            join contexto using (cod_contexto) 
                            join entidad using (cod_entidad)
                            join periodo using (cod_periodo) 
                            left join context_escenarios using (cod_contexto) 
                            left join escenario using (cod_escenario) 
                            left join dimension using (cod_dimension) 
                            --where cod_periodo = 44 and cod_tipo_cuenta in (6,7)
                            where cod_dimension is null  and cod_documento = %s and cod_role =%s
                            order by cod_role_cuenta, desc_cuenta""" % (cod_documento, cod_role))
    data = dictfetchall(cursor)
    cursor.execute("""select desc_role, cod_role_cuenta,desc_escenario, desc_dimension, desc_tipo_cuenta,coalesce(desc_role_cuenta, etiqueta, desc_cuenta) desc_cuenta, cod_periodo,desc_periodo, desc_detalle_documento, decimales, cod_unidad, desc_unidad, cod_entidad,desc_entidad, case when not decimales isnull then desc_detalle_documento::numeric / 10 ^ abs(decimales) else 0 end valor 
                                from detalle_documento 
                                join cuenta using (cod_cuenta) 
                                left join role_cuenta using (cod_cuenta)
                                left join role using(cod_role)
                                left join unidad using (cod_unidad)
                                join tipo_cuenta using (cod_tipo_cuenta)
                                join contexto using (cod_contexto) 
                                join entidad using (cod_entidad)
                                join periodo using (cod_periodo) 
                                left join context_escenarios using (cod_contexto) 
                                left join escenario using (cod_escenario) 
                                left join dimension using (cod_dimension) 
                                --where cod_periodo = 44 and cod_tipo_cuenta in (6,7)
                                where cod_dimension is null  and cod_documento = %s and cod_role =%s
                                order by cod_role_cuenta, desc_cuenta""" % (cod_documento, cod_role))
    date = cursor.fetchall()
    cursor.execute("""select distinct desc_periodo 
                               from detalle_documento 
                               join cuenta using (cod_cuenta) 
                               left join role_cuenta using (cod_cuenta)
                               left join role using(cod_role)
                               join contexto using (cod_contexto) 
                               join periodo using (cod_periodo)
                               left join context_escenarios using (cod_contexto) 
                               left join escenario using (cod_escenario) 
                               left join dimension using (cod_dimension)
                               where cod_dimension is null  and cod_documento = %s and cod_role =%s
                               order by desc_periodo desc""" % (cod_documento, cod_role))
    per = dictfetchall(cursor)
    for d in date:
        cod_rol = d[1]
        desc_role = d[0]
        entidad = d[13]
        cuenta = d[5]
        dineros = d[8]
        fechas = d[7]
    hoy = datetime.now()
    hoy = hoy.strftime("%Y-%m-%d")
    report = PdfCustomDetail(filename="%s.pdf" % (cod_rol), title=["%s" % (desc_role), entidad.__str__(), cuenta],
                             logo=getattr(sys, 'logo', 'xeeffy_logo_index.png'))
    pvt_kms = pivot(data, ('cod_role_cuenta', 'desc_cuenta'), ('desc_periodo',), 'valor')
    datas = []
    columnas = ["CUENTA"]
    data_linea = ['']
    ancho_total = 0
    colsWidth = [0]
    colAlign = ['LEFT']
    colType = ['str']
    for p in per:
        columnas.append(p['desc_periodo'])
        data_linea.append(0.0)
        colsWidth.append(60)
        ancho_total += 60
        colAlign.append("RIGHT")
        colType.append("str")
    colsWidth[0] = 540 - ancho_total
    for f in pvt_kms:
        dl = list(f.get(('cod_role_cuenta', 'desc_cuenta'), ['']))
        for p in per:
            dl.append(f.get((p['desc_periodo'],), 0))
        datas.append(dl)
    data_final = []
    datas.sort()
    for d in datas:
        data_final.append(d[1:])
    return render_template("/pdf.html", d=data_final, columnas=columnas, entidad=entidad, desc_role=desc_role, hoy=hoy)
    # report.drawData(data_final, columnas, colswidht=colsWidth, fontSize=7, ColsAlign=colAlign, ColsType=colType)
    # report.go(tmp=True)

    # return render_template("/pdf.html", desc_role=desc_role, entidad=entidad,
    #                        cuenta=cuenta,
    #                        dineros=dineros, fechas=fechas, hoy=hoy, data=data)
    # fileNameOutput = 'informe.pdf'
    # # css = ['/static/css/pdf.css']
    # pdfkit.from_string(cont, fileNameOutput)
    # pdfDownload = open(fileNameOutput, 'rb').read()
    # os.remove(fileNameOutput)
    # return Response(
    #     pdfDownload,
    #     mimetype="application/pdf",
    #     headers={
    #         "Content-disposition": "attachment; filename=" + fileNameOutput,
    #         "Content-type": "application/force-download"
    #     }
    # )
    # return render_template("/documentos.html")


def __build_dict(description, row):
    res = {}
    for i in range(len(description)):
        res[description[i][0]] = row[i]
    return res


def dictfetchall(cursor):
    res = []
    rows = cursor.fetchall()
    for row in rows:
        res.append(__build_dict(cursor.description, row))
    return res


class PdfCustomDetail:
    def __init__(self, title="Documento Generico", filename="output.pdf", barcode=None, companyInfo=[], logo=None,
                 distintoEncabezado=True):
        self.Elements = []
        self.Data = []
        self.Headers = []
        self.logo = logo
        self.filename = filename.replace(" ", "_")
        self.distintoEncabezado = distintoEncabezado

        self.styleSheet = getSampleStyleSheet()

        self.Title = title
        self.Author = "XSolution ERP"

        self.URL = "http://www.xsolution.cl"
        if companyInfo == []:
            self.CompanyInfo = [["Empresa.", 0],
                                ["RUT: 99.000.000-0", 1],
                                ["Direccion", 2],
                                ["Fonos", 3]]
        else:
            self.CompanyInfo = companyInfo
        self.email = "contacto@xsolution.cl"
        self.pageinfo = "%s / %s / %s" % (self.Author, self.email, self.Title)
        self.barcode = barcode


@app.route('/importCMF', methods=["GET", "POST"])
def importCMF():
    if session["username"] != 'unknown':
        if request.method == "POST":
            import urllib.request
            import urllib

            cur = cnx.cursor()
            ano = request.form["ano"]
            mes = request.form["mes"]
            mes = mes.zfill(2)
            url = r"http://www.cmfchile.cl/institucional/mercados/novedades_envio_sa_ifrs_excel2.php?aa=%s&mm=%s" % (
                ano, mes)
            filename = "%s%s%s_%s%s.xls" % (
                tempfile._get_default_tempdir(), os.sep, tempfile._RandomNameSequence().__next__(), ano, mes)
            opener = urllib.request.build_opener()
            opener.addheaders = [('User-Agent',
                                  'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1941.0 Safari/537.36')]
            user_agent = 'Mozilla/5.0 (Windows NT 6.1; Win64; x64)'
            values = {'name': 'Michael Foord',
                      'location': 'Northampton',
                      'language': 'Python'}
            headers = {'User-Agent': user_agent}
            urllib.request.install_opener(opener)
            data1 = urllib.request.urlretrieve(url, filename)
            http = urllib3.PoolManager()
            response = http.request('GET', url)
            soup = BeautifulSoup(response.data)
            book = xlrd.open_workbook(filename)
            sh = book.sheet_by_index(0)
            for rx in range(4, sh.nrows, 1):
                fpenvio = sh.cell(rx, 0).value
                fuenvio = sh.cell(rx, 1).value
                razon_social = sh.cell(rx, 2).value
                tipo_balance = sh.cell(rx, 3).value
                rut, dv = sh.cell(rx, 4).value.split("-")
                cur.execute("""
                SELECT
                public.entidad.cod_entidad,
                public.entidad.desc_entidad,
                public.entidad.user_id
                FROM
                public.entidad
                WHERE entidad.cod_entidad = '%s'""" % (CRut("%s-%s" % (rut, dv))))
                dato = cur.fetchall()
                if not dato:
                    flash("No existe")
                    cur.execute("""INSERT INTO entidad(cod_entidad,desc_entidad)
                    VALUES('%s','%s')""" % (CRut("%s-%s" % (rut, dv)), razon_social))
                    cnx.commit()
                tipo_envio = sh.cell(rx, 5).value
                filtros = {'rut': rut, 'mes': mes, 'anio': ano, 'archivo': "Estados%20financieros%20(XBRL)",
                           'tipo': tipo_balance.upper()[0]}
                desc_doc = "%(rut)s_%(anio)s%(mes)s_%(tipo)s.zip" % (filtros)
                cur.execute("""SELECT
                public.documento.cod_documento,
                public.documento.desc_documento,
                public.documento.cod_entidad,
                public.entidad.desc_entidad,
                public.documento.anio,
                public.documento.mes
                FROM
                public.documento,
                public.entidad
                WHERE documento.cod_entidad='%s' and documento.desc_documento ='%s' and entidad.desc_entidad = '%s' and documento.anio='%s' and documento.mes ='%s'""" % (
                    CRut("%s-%s" % (rut, dv)), desc_doc, razon_social, ano, mes))
                doc = cur.fetchall()
                if not doc:
                    print("NUEVO!!!", desc_doc, rx, "de", sh.nrows)
                    has = "*"
                    cur.execute(
                        """INSERT INTO documento(desc_documento,cod_entidad,anio,mes,hash)VALUES('%s','%s','%s','%s','%s')""" % (
                            desc_doc, CRut("%s-%s" % (rut, dv)), ano, mes, has))

                    url_zip = "http://www.cmfchile.cl/institucional/inc/inf_financiera/ifrs/safec_ifrs_verarchivo.php?auth=&send=&rut=%(rut)s&mm=%(mes)s&aa=%(anio)s&archivo=%(rut)s_%(anio)s%(mes)s_%(tipo)s.zip&desc_archivo=%(archivo)s&tipo_archivo=XBRL" % (
                        filtros)
                    filename_zip = tempfile._get_default_tempdir() + os.sep + tempfile._RandomNameSequence().__next__() + "%(rut)s_%(anio)s%(mes)s_C.zip" % (
                        filtros)
                    try:
                        values = {'name': 'Michael Foord',
                                  'location': 'Northampton',
                                  'language': 'Python'}
                        data1 = urllib.parse.urlencode(values).encode("utf-8")
                        req = urllib.request.Request(url_zip, data1, headers)
                        with urllib.request.urlopen(req, data=data1) as f:
                            resp = f.read()
                        f = open(filename_zip, "wb")
                        f.write(resp)
                        f.close()
                        input_zip = zipfile.ZipFile(filename_zip)
                        data = {name: input_zip.read(name) for name in input_zip.namelist() if
                                name.upper().endswith('.XBRL')}
                        for arch in data.keys():
                            # jxbrl = jsobXBrl(cod_documento=doc.get_cod_documento())
                            cursor = cnx.cursor()
                            cursor.execute("SELECT currval('documento_cod_documento_seq')")
                            jxbrl = jsobXBrl(cod_documento=cursor.fetchone()[0])
                            jxbrl.loadXBrlFromMenory(data.get(arch))
                            jxbrl.processJsonXbrl()
                            print("hola")
                        #     del jxbrl
                        # del data
                        cnx.commit()
                    except:
                        print("ERROR", desc_doc, sys.exc_info()[1])
                else:
                    print("EXISTE!!!", desc_doc, rx, "de", sh.nrows)
                del doc
            return redirect("/documentos")
        return render_template("/importCMF.html")
    return redirect("/login")


class jsobXBrl():
    def __init__(self, xbrlFile=None, cod_documento=None):
        self.cod_documento = cod_documento
        self.xbrlFile = xbrlFile
        self.jsonXBrl = None
        self.contexts = {}
        self.data = []
        self.entity = {}
        self.period = {}
        self.scenario = {}
        self.tipo_scenario = {}
        self.dimension = {}
        self.tipoCuenta = {}
        self.cuentas = {}
        self.unidades = {}
        self.loadUnidades()
        self.loadTipoCuenta()
        self.loadCuentas()
        self.loadEntity()
        self.loadPeriod()
        self.loadDimension()
        self.loadTipoScenario()

        pass

    def processJsonXbrl(self):
        for i in self.jsonXBrl:
            if i.split(":")[-1] == 'context':
                for c in self.jsonXBrl.get(i):
                    self.addContext(c)
            elif i.split(":")[-1] == 'unit':
                for u in self.jsonXBrl.get(i):
                    self.addUnit(u)
            else:
                pass
        for i in self.jsonXBrl:
            if i == 'context':
                pass
            elif i == 'unit':
                pass
            else:
                self.addData(i, self.jsonXBrl.get(i))
        print("OK")

    def addCuenta(self, data):
        # from orm.common_services.public.cms_cuenta import cms_cuenta
        # from orm.common_services.public.cms_tipo_cuenta import cms_tipo_cuenta
        tc, cta = data.split(":")
        ccta = self.cuentas.get(":".join([{'ifrs': "ifrs-full"}.get(tc, tc), cta]))

        if ccta:
            return ccta
        if not self.tipoCuenta.get(tc):
            cursor = cnx.cursor()
            cursor.execute("insert into tipo_cuenta(desc_tipo_cuenta) values ('%s')"%(tc))
            # tcta = cms_tipo_cuenta()
            # tcta.set_desc_tipo_cuenta(tc)
            cnx.commit()
            # tcta.save()
            cursor.execute("SELECT currval('tipo_cuenta_cod_tipo_cuenta_seq')")
            # ctc = tcta.get_cod_tipo_cuenta()
            self.tipoCuenta[tc] = cursor.fetchone()[0]
        else:
            ctc = self.tipoCuenta.get(tc)
        cursor = cnx.cursor()
        cursor.execute("insert into cuenta(cod_tipo_cuenta,desc_cuenta) values ('%s','%s')"%(ctc,cta))
        cnx.commit()
        cursor.execute("SELECT currval('cuenta_cod_cuenta_seq')")
        # objcta = cms_cuenta()
        # objcta.set_desc_cuenta(cta)
        # objcta.set_cod_tipo_cuenta(ctc)
        # objcta.save()
        # ccta = objcta.get_cod_cuenta()
        ccta = cursor.fetchone()[0]
        self.cuentas[data] = ccta
        return ccta

    def loadTipoCuenta(self):
        #from orm.common_services.public.cms_tipo_cuenta import cms_tipo_cuenta
        #tc = cms_tipo_cuenta()
        cursor = cnx.cursor()
        cursor.execute("select cod_tipo_cuenta, desc_tipo_cuenta from tipo_cuenta")
        for t in cursor.fetchall():
            # self.tipoCuenta[t.get_desc_tipo_cuenta()] = t.get_cod_tipo_cuenta()
            self.tipoCuenta[t[1]] = t[0]

    def loadCuentas(self):
        #from orm.common_services.public.cms_cuenta import cms_cuenta
        #cs = cms_cuenta()
        cursor = cnx.cursor()
        cursor.execute("""SELECT
        public.tipo_cuenta.desc_tipo_cuenta,
        public.cuenta.desc_cuenta,
        public.cuenta.cod_cuenta,
        public.cuenta.cod_tipo_cuenta
        FROM
        public.cuenta
        INNER JOIN public.tipo_cuenta ON public.cuenta.cod_tipo_cuenta = public.tipo_cuenta.cod_tipo_cuenta""")
        #for c in cs.custom_query(
                # "select desc_tipo_cuenta||':'||desc_cuenta, cod_cuenta from cuenta join tipo_cuenta using (cod_tipo_cuenta)"):
        for c in cursor.fetchall():
            self.cuentas[c[0]] = c[1]

    def addData(self, t, data):
        try:
            cursor = cnx.cursor()
            cta = self.cuentas.get(t)
            if not cta:
                cta = self.addCuenta(t)

            if isinstance(data, dict):
                c = self.contexts.get(data.get("@contextRef"), {}).get('id')
                d = data.get('@decimals')
                u = self.unidades.get(data.get('@unitRef'))
                valor = data.get("#text")
                cursor.execute(
                    """insert into detalle_documento (cod_documento, cod_cuenta, cod_contexto ,decimales, cod_unidad,desc_detalle_documento)values('%s','%s','%s','%s','%s','%s')""" % (
                        self.cod_documento, cta, c, d, u, valor))
                # dd = cms_detalle_documento()
                # dd.set_cod_documento(self.cod_documento)
                # dd.set_cod_cuenta(cta)
                # dd.set_cod_contexto(c)
                # dd.set_decimales(d)
                # dd.set_cod_unidad(u)
                # dd.set_desc_detalle_documento(valor)
                # dd.save()
                cnx.commit()
                self.data.append([cta, c, d, u, valor])
            elif isinstance(data, list):
                for i in data:
                    c = self.contexts.get(i.get("@contextRef"), {}).get('id')
                    d = i.get('@decimals')
                    u = self.unidades.get(i.get('@unitRef'))
                    valor = i.get("#text")
                    cursor.execute(
                        """insert into detalle_documento (cod_documento, cod_cuenta, cod_contexto ,decimales, cod_unidad,desc_detalle_documento)values('%s','%s','%s','%s','%s','%s')""" % (
                        self.cod_documento, cta, c, d, u, valor))
                    # dd = cms_detalle_documento()
                    # dd.set_cod_documento(self.cod_documento)
                    # dd.set_cod_cuenta(cta)
                    # dd.set_cod_contexto(c)
                    # dd.set_decimales(d)
                    # dd.set_cod_unidad(u)
                    # dd.set_desc_detalle_documento(valor)
                    # dd.save()
                    cnx.commit()
                    self.data.append([cta, c, d, u, valor])
            else:
                print(t, data)
        except:
            print(t, data)
        # print len(self.data)

    def loadUnidades(self):
        #from orm.common_services.public.cms_unidad import cms_unidad
        cursor = cnx.cursor()
        cursor.execute("select cod_unidad,desc_unidad from unidad")
        for u in cursor.fetchall():
            # self.unidades[u.get_desc_unidad()] = u.get_cod_unidad()
            self.unidades[u[1]] = u[0]

    def addUnit(self, data):
        cursor = cnx.cursor()
        # from orm.common_services.public.cms_unidad import cms_unidad
        if self.unidades.get(data.get("@id")):
            return
        # uni = cms_unidad()
        # uni.set_desc_unidad(data.get("@id"))
        # uni.save()
        cursor.execute("insert into unidad(desc_unidad)values ('%s')" % (data.get("@id")))
        cnx.commit()
        # cursor.execute("select cod_unidad from unidad where desc_unidad ='%s'"% (data.get("@id")))
        # uni = cursor.fetchall()
        cursor.execute("SELECT currval('unidad_cod_unidad_seq')")
        self.unidades[data.get("@id")] = cursor.fetchone()[0]
        print(data)

    def addDimension(self, data):
        cod = self.dimension.get(data.get('@dimension'))
        if not cod:
            # from orm.common_services.public.cms_dimension import cms_dimension
            cursor = cnx.cursor()
            cursor.execute("insert into dimension(desc_dimension)values('%s')" % (data.get('@dimension')))
            cnx.commit()
            # cursor.execute("select cod_dimension from dimension where desc_dimension = '%s'" % (data.get('@dimension')))
            # cod = cursor.fetchall()
            cursor.execute("SELECT currval('dimension_cod_dimension_seq')")
            # obj = cms_dimension()
            # obj.set_desc_dimension(data.get('@dimension'))
            # obj.save()
            # cod = obj.get_cod_dimension()
            self.dimension[data.get('@dimension')] = cursor.fetchone()[0]
        return cod

    def loadDimension(self):
        #from orm.common_services.public.cms_dimension import cms_dimension
        # obj = cms_dimension()
        cursor = cnx.cursor()
        cursor.execute("select cod_dimension , desc_dimension from dimension")
        for obj in cursor.fetchall():
            #self.dimension[obj.get_desc_dimension()] = obj.get_cod_dimension()
            self.dimension[obj[1]] = obj[0]

    def loadScenario(self):
        #from orm.common_services.public.cms_escenario import cms_escenario
        cursor = cnx.cursor()
        cursor.execute("""SELECT
        public.escenario.desc_escenario,
        public.tipo_escenario.desc_tipo_escenario,
        public.dimension.desc_dimension,
        public.escenario.cod_escenario
        FROM
        public.escenario
        INNER JOIN public.tipo_escenario ON public.escenario.cod_tipo_escenario = public.tipo_escenario.cod_tipo_escenario
        INNER JOIN public.dimension ON public.escenario.cod_dimension = public.dimension.cod_dimension""")
        #es = cms_escenario()
        for es in cursor.fetchall():
            # self.scenario[(es.get_desc_escenario(), es.get_tipo_escenario().__str__(),
            #                es.get_dimension().__str__())] = es.get_cod_escenario()
            self.scenario[es[0],es[1],es[2]] = es[3]

    def addScenario(self, data):
        cods = []
        if data:
            for key in data.keys():
                if isinstance(data.get(key), dict):
                    escenarios = [data.get(key)]
                else:
                    escenarios = data.get(key)
                for escenario in escenarios:
                    cod = self.scenario.get((escenario.get('#text'), key, escenario.get('@dimension')))
                    if not cod:
                        cursor = cnx.cursor()
                        ts = self.addTipoScenario(key)
                        d = self.addDimension(escenario)
                        #from orm.common_services.public.cms_escenario import cms_escenario
                        cursor.execute(
                            "insert into escenario(desc_escenario,cod_dimension,cod_tipo_escenario)values('%s','%s','%s')" % (
                                escenario.get('#text'), d, ts))
                        cursor.execute("SELECT currval('escenario_cod_escenario_seq')")
                        cnx.commit()
                        cod = cursor.fetchone()[0]
                        self.scenario[(escenario.get('#text'), key, escenario.get('@dimension'))] = cod
                    cods.append(cod)
        return cods

    def addTipoScenario(self, key):
        cod = self.tipo_scenario.get(key)
        if not cod:
            # from orm.common_services.public.cms_tipo_escenario import cms_tipo_escenario as clas
            cursor = cnx.cursor()
            cursor.execute("insert into tipo_escenario(desc_tipo escenario)values('%s')" % (key))
            cnx.commit()
            # obj = clas()
            # obj.set_desc_tipo_escenario(key)
            # obj.save()
            cursor.execute("SELECT currval('tipo_escenario_cod_tipo_escenario_seq')")
            # cursor.execute("select cod_tipo_escenario from tipo_escenario where desc_tipo_escenario='%s'" % (key))
            # cod = obj.get_cod_tipo_escenario()
            # cod = cursor.fetchall()
            self.tipo_scenario[key] = cursor.fetchone()[0]
        return cod

    def loadTipoScenario(self):
        #from orm.common_services.public.cms_tipo_escenario import cms_tipo_escenario as clas
        # obj = clas()
        cursor = cnx.cursor()
        cursor.execute("select cod_tipo_escenario,desc_tipo_escenario from tipo_escenario")
        for obj in cursor.fetchall():
            # self.tipo_scenario[obj.get_desc_tipo_escenario()] = obj.get_cod_tipo_escenario()
            self.tipo_scenario[obj[1]] = obj[0]

    def loadXBrlFromFile(self):
        md5 = self.hash_bytestr_iter(self.file_as_blockiter(open(self.xbrlFile, 'rb')), hashlib.sha256(), True)
        fxml = open(self.xbrlFile)
        xml = fxml.read()
        xml = eval(xmltodict.parse(xml))
        self.jsonXBrl = None
        keys = ['xbrli:xbrl', 'xbrl']
        for key in keys:
            self.jsonXBrl = xml.get(key)
            if self.jsonXBrl:
                return
        print("No Encontro XBRL")

    def hash_bytestr_iter(self, bytesiter, hasher, ashexstr=False):
        for block in bytesiter:
            hasher.update(block)
        return (hasher.hexdigest() if ashexstr else hasher.digest())

    def file_as_blockiter(self, afile, blocksize=65536):
        with afile:
            block = afile.read(blocksize)
            while len(block) > 0:
                yield block
                block = afile.read(blocksize)

    def loadEntity(self):
        #from orm.common_services.public.cms_entidad import cms_entidad
        cursor = cnx.cursor()
        cursor.execute("select cod_entidad from entidad")
        for e in cursor.fetchall():
            #self.entity[e.get_cod_entidad()] = e.get_cod_entidad()
            self.entity[e[0]] = e[0]

    def addEntity(self, data):
        cod = self.entity.get(CRut(data.get('identifier', data.get('xbrli:identifier')).get('#text')))
        cod1 = cod
        if not cod:
            cursor = cnx.cursor()
            #from orm.common_services.public.cms_entidad import cms_entidad
            cursor.execute("insert into entidad (cod_entidad, desc_entidad) values ('%s','%s')" % (CRut(data.get('identifier', data.get('xbrli:identifier')).get('#text')),data.get('identifier', data.get('xbrli:identifier')).get('#text')))
            cnx.commit()
            # enty = cms_entidad()
            # enty.set_cod_entidad((CRut(data.get('identifier', data.get('xbrli:identifier')).get('#text'))))
            # enty.set_desc_entidad(data.get('identifier', data.get('xbrli:identifier')).get('#text'))
            # enty.save()
            cursor.execute("select cod_entidad from entidad where cod_entidad ='%s' and desc_entidad ='%s'" % (
            (CRut(data.get('identifier', data.get('xbrli:identifier')).get('#text')))),
                           data.get('identifier', data.get('xbrli:identifier')).get('#text'))
            cod = cursor.fetchall()
            # cod = enty.get_cod_entidad()
            # self.entity[enty.get_cod_entidad()] = cod
            self.entity[CRut(data.get('identifier', data.get('xbrli:identifier')).get('#text'))] = cod[0][0]
        return cod

    def addContext(self, data):
        cursor = cnx.cursor()
        if [k for k in data.keys() if not k.split(":")[-1] in ("entity", 'period', 'scenario', '@id')]:
            print(data)
        if self.contexts.get(data['@id']):
            return
        e = self.addEntity(data.get("entity", data.get("xbrli:entity")))
        self.addPeriod(data.get('period', data.get('xbrli:period')))
        p = self.getPeriod(data.get('period', data.get('xbrli:period')))
        scenarios = self.addScenario(data.get('scenario', data.get('xbrli:scenario')))
        #scenarios = self.getScenario(data.get('scenario',data.get('xbrli:scenario')))
        # from orm.common_services.public.cms_contexto import cms_contexto
        # c = cms_contexto()
        # c.set_cod_entidad(e)
        # c.set_cod_periodo(p)
        # # c.set_cod_escenario(s)
        # c.set_desc_contexto(data['@id'])
        # c.save()
        cursor.execute(
            "insert into contexto(cod_entidad, cod_periodo, desc_contexto) values ('%s','%s','%s')" % (e, p, data['@id']))
        cnx.commit()
        cursor.execute("SELECT currval('contexto_cod_contexto_seq')")
        codigo_contexto = cursor.fetchone()[0]
        self.contexts[data['@id']] = {"entity": e, "period": p, "id": codigo_contexto}
        #from orm.common_services.public.cms_context_escenarios import cms_context_escenarios
        for s in scenarios:

            cursor.execute("insert into context_escenarios(cod_contexto,cod_escenario,desc_context_escenarios) values('%s','%s','.')"%(codigo_contexto,s))
            # ce = cms_context_escenarios()
            # ce.set_cod_contexto(c.get_cod_contexto())
            # ce.set_cod_escenario(s)
            # ce.set_desc_context_escenarios(".")
            # ce.save()
            cnx.commit()

    def getScenario(self, data):
        if data:
            for key in data.keys():
                return self.scenario[(data.get(key).get('#text'), key, data.get(key).get('@dimension'))]

        # def getScenario(self, data):
        #     if data:
        #         for key in data.keys():
        #             return self.scenario[(data.get(key).get('#text'), key, data.get(key).get('@dimension'))]
        # cursor.execute("select cod_contexto from contexto where cod_entidad='%s'and cod_periodo='%s' and desc_contexto = '%s'"%(e, p, data['@id']))
        # a = cursor.fetchall()
        # self.contexts[data['@id']] = {"entity": e, "period": p, "id": a}
        #self.contexts[data['@id']] = {"entity": e, "period": p, "id": c.get_cod_contexto()}
        # from orm.common_services.public.cms_context_escenarios import cms_context_escenarios
        # for s in scenarios:
        #     cursor.execute = (
        #             "insert into context_excenarios(cod_contexto,cod_escenario,desc_context_escenario) values ('%s','%s','%s')" % (a, s, "."))
        #     f = cursor.fetchall()
        #     cnx.commit()
            # ce = cms_context_escenarios()
            # ce.set_cod_contexto(c.get_cod_contexto())
            # ce.set_cod_escenario(s)
            # ce.set_desc_context_escenarios(".")
            # ce.save()

    def getPeriod(self, data):
        return self.period.get((data.get('startDate', data.get('xbrli:startDate')),
                                data.get('endDate', data.get('xbrli:endDate')),
                                data.get("instant", data.get("xbrli:instant"))))

    def addPeriod(self, data):
        cursor =cnx.cursor()
        if not self.period.get((data.get('startDate', data.get('xbrli:startDate')),
                                data.get('endDate', data.get('xbrli:endDate')),
                                data.get("instant", data.get("xbrli:instant")))):
            #from orm.common_services.public.cms_periodo import cms_periodo
            sql = "insert into periodo(desc_periodo,"
            valores = "('.',"
            for d in data:
                if data[d]:
                    sql += d.replace('xbrli:','')+ ","
                    valores += "'"+data[d] + "',"
            sql=sql[:-1]+ ") values "+ valores [:-1] + ")"
            cursor.execute(sql)
            # if data.get("xbrli:instant"):
            #     cursor.execute("insert into periodo (startdate,enddate,instant,desc_periodo)values ('%s','%s','%s','.')"%(data.get('startDate', data.get('xbrli:startDate')),data.get('endDate', data.get('xbrli:endDate')),data.get('instant', data.get('xbrli:instant'))or 'Null'))
            # else:
            #     cursor.execute("insert into periodo (startdate,enddate,desc_periodo)values ('%s','%s','.')" % (
            #     data.get('startDate', data.get('xbrli:startDate')), data.get('endDate', data.get('xbrli:endDate'))
            #     ))

            cnx.commit()
            cursor.execute("SELECT currval ('periodo_cod_periodo_seq')")
            self.period[(
                data.get('startDate', data.get('xbrli:startDate')), data.get('endDate', data.get('xbrli:endDate')),
                data.get("instant", data.get("xbrli:instant")))] = cursor.fetchone()[0]
            # p = cms_periodo()
            # p.set_startdate(data.get('startDate', data.get('xbrli:startDate')))
            # p.set_enddate(data.get('endDate', data.get('xbrli:endDate')))
            # p.set_instant(data.get('instant', data.get('xbrli:instant')))
            # p.set_desc_periodo(".")
            # p._pre_save()
            # p.save()
            # self.period[(
            #     data.get('startDate', data.get('xbrli:startDate')), data.get('endDate', data.get('xbrli:endDate')),
            #     data.get("instant", data.get("xbrli:instant")))] = p.get_cod_periodo()


    def loadPeriod(self):
        #from orm.common_services.public.cms_periodo import cms_periodo
        cursor = cnx.cursor()
        cursor.execute("select cod_periodo, startdate, enddate,instant from periodo")
        # per = cms_periodo()
        for p in cursor.fetchall():
            # s = p.get_startdate().__str__() if p.get_startdate() else None
            # e = p.get_enddate().__str__() if p.get_enddate() else None
            # i = p.get_instant().__str__() if p.get_instant() else None
            self.period[(p[1], p[2], p[3])] = p[0]


    def loadXBrlFromMenory(self, dataXML):
        xbrl_parser = XBRLParser()
        soup = BeautifulSoup(dataXML, 'lxml')
        btree = BeautifulSoup(dataXML, 'lxml')
        Terms = btree.select('xbrli > xbrli')
        jsonObj = {"xbrli": []}
        for term in Terms:
            termDetail = {
                "xbrli:xbrl": term.find('xbrli:xbrl').text,
                "xbrl": term.find('xbrl').text
            }
            RelatedTerms = term.select('RelatedTerms > Term')
            if RelatedTerms:
                termDetail["RelatedTerms"] = []
                for rterm in RelatedTerms:
                    termDetail["RelatedTerms"].append({
                        "Title": rterm.find('Title').text,
                        "Relationship": rterm.find('Relationship').text
                    })
            jsonObj["thesaurus"].append(termDetail)
        tag_list = soup.find_all()
        for tag in tag_list:
            if tag.name == 'us-gaap:liabilities':
                print('Liabilities: ' + tag.text)
        import json
        import xmltodict
        # xml = eval(xmljson(dataXML).replace("null", "None"))
        jsonString = xmltodict.parse(dataXML)

        self.jsonXBrl = None
        keys = ['xbrli:xbrl', 'xbrl']
        for key in keys:
            self.jsonXBrl = jsonString.get(key)
            if self.jsonXBrl:
                return
        print("No Encontro XBRL")

    def loadContext(self):
        #from orm.common_services.public.cms_contexto import cms_contexto
        #c = cms_contexto()
        cursor = cnx.cursor()
        cursor.execute("select cod_contexto,cod_entidad,cod_periodo,desc_contexto from cotexto")
        for c in cursor.fetchall():
            # self.contexts[c.get_desc_contexto()] = {"entity": c.get_cod_entidad(), "period": c.get_cod_periodo(),
            #                                         "id": c.get_cod_contexto()}
            self.contexts[c[3]] = {"entity": c[1], "period": c[2], "id": c[0]}



@app.route('/eresultado/<cod_documento>/<cod_role>', methods=["GET", "POST"])
def eresultado(cod_documento, cod_role):
    cursor = cnx.cursor()
    cursor.execute("""select desc_role, cod_role_cuenta,desc_escenario, desc_dimension, desc_tipo_cuenta,coalesce(desc_role_cuenta, etiqueta, desc_cuenta) desc_cuenta, cod_periodo,desc_periodo, desc_detalle_documento, decimales, cod_unidad, desc_unidad, cod_entidad,desc_entidad, case when not decimales isnull then desc_detalle_documento::numeric / 10 ^ abs(decimales) else 0 end valor 
                            from detalle_documento 
                            join cuenta using (cod_cuenta) 
                            left join role_cuenta using (cod_cuenta)
                            left join role using(cod_role)
                            left join unidad using (cod_unidad)
                            join tipo_cuenta using (cod_tipo_cuenta)
                            join contexto using (cod_contexto) 
                            join entidad using (cod_entidad)
                            join periodo using (cod_periodo) 
                            left join context_escenarios using (cod_contexto) 
                            left join escenario using (cod_escenario) 
                            left join dimension using (cod_dimension) 
                            --where cod_periodo = 44 and cod_tipo_cuenta in (6,7)
                            where cod_dimension is null  and cod_documento = %s and cod_role =%s
                            order by cod_role_cuenta, desc_cuenta""" % (cod_documento, cod_role))
    data = dictfetchall(cursor)
    cursor.execute("""select desc_role, cod_role_cuenta,desc_escenario, desc_dimension, desc_tipo_cuenta,coalesce(desc_role_cuenta, etiqueta, desc_cuenta) desc_cuenta, cod_periodo,desc_periodo, desc_detalle_documento, decimales, cod_unidad, desc_unidad, cod_entidad,desc_entidad, case when not decimales isnull then desc_detalle_documento::numeric / 10 ^ abs(decimales) else 0 end valor 
                                from detalle_documento 
                                join cuenta using (cod_cuenta) 
                                left join role_cuenta using (cod_cuenta)
                                left join role using(cod_role)
                                left join unidad using (cod_unidad)
                                join tipo_cuenta using (cod_tipo_cuenta)
                                join contexto using (cod_contexto) 
                                join entidad using (cod_entidad)
                                join periodo using (cod_periodo) 
                                left join context_escenarios using (cod_contexto) 
                                left join escenario using (cod_escenario) 
                                left join dimension using (cod_dimension) 
                                --where cod_periodo = 44 and cod_tipo_cuenta in (6,7)
                                where cod_dimension is null  and cod_documento = %s and cod_role =%s
                                order by cod_role_cuenta, desc_cuenta""" % (cod_documento, cod_role))
    date = cursor.fetchall()
    cursor.execute("""select distinct desc_periodo 
                               from detalle_documento 
                               join cuenta using (cod_cuenta) 
                               left join role_cuenta using (cod_cuenta)
                               left join role using(cod_role)
                               join contexto using (cod_contexto) 
                               join periodo using (cod_periodo)
                               left join context_escenarios using (cod_contexto) 
                               left join escenario using (cod_escenario) 
                               left join dimension using (cod_dimension)
                               where cod_dimension is null  and cod_documento = %s and cod_role =%s
                               order by desc_periodo desc""" % (cod_documento, cod_role))
    per = dictfetchall(cursor)
    for d in date:
        cod_rol = d[1]
        desc_role = d[0]
        entidad = d[13]
        cuenta = d[5]
        dineros = d[8]
        fechas = d[7]
    hoy = datetime.now()
    hoy = hoy.strftime("%Y-%m-%d")
    report = PdfCustomDetail(filename="%s.pdf" % (cod_rol), title=["%s" % (desc_role), entidad.__str__(), cuenta],
                             logo=getattr(sys, 'logo', 'xeeffy_logo_index.png'))
    pvt_kms = pivot(data, ('cod_role_cuenta', 'desc_cuenta'), ('desc_periodo',), 'valor')
    datas = []
    columnas = ["CUENTA"]
    data_linea = ['']
    ancho_total = 0
    colsWidth = [0]
    colAlign = ['LEFT']
    colType = ['str']
    for p in per:
        columnas.append(p['desc_periodo'])
        data_linea.append(0.0)
        colsWidth.append(60)
        ancho_total += 60
        colAlign.append("RIGHT")
        colType.append("str")
    colsWidth[0] = 540 - ancho_total
    for f in pvt_kms:
        dl = list(f.get(('cod_role_cuenta', 'desc_cuenta'), ['']))
        for p in per:
            dl.append(f.get((p['desc_periodo'],), 0))
        datas.append(dl)
    data_final = []
    datas.sort()
    for d in datas:
        data_final.append(d[1:])
    return render_template("/eresultado.html", d=data_final, columnas=columnas, entidad=entidad, desc_role=desc_role,
                           hoy=hoy)


@app.route('/ratioLiquidez/<currentA>/<currentL>', methods=["GET", "POST"])
def ratioLiquidez(currentA, currentL):
    return render_template("/ratioLiquidez.html")


@app.route('/balance/<cod_documento>/<cod_role>', methods=["GET", "POST"])
def balance(cod_documento, cod_role):
    cursor = cnx.cursor()
    cursor.execute("""select desc_role, cod_role_cuenta,desc_escenario, desc_dimension, desc_tipo_cuenta,coalesce(desc_role_cuenta, etiqueta, desc_cuenta) desc_cuenta, cod_periodo,desc_periodo, desc_detalle_documento, decimales, cod_unidad, desc_unidad, cod_entidad,desc_entidad, case when not decimales isnull then desc_detalle_documento::numeric / 10 ^ abs(decimales) else 0 end valor 
                            from detalle_documento 
                            join cuenta using (cod_cuenta) 
                            left join role_cuenta using (cod_cuenta)
                            left join role using(cod_role)
                            left join unidad using (cod_unidad)
                            join tipo_cuenta using (cod_tipo_cuenta)
                            join contexto using (cod_contexto) 
                            join entidad using (cod_entidad)
                            join periodo using (cod_periodo) 
                            left join context_escenarios using (cod_contexto) 
                            left join escenario using (cod_escenario) 
                            left join dimension using (cod_dimension) 
                            --where cod_periodo = 44 and cod_tipo_cuenta in (6,7)
                            where cod_dimension is null  and cod_documento = %s and cod_role =%s
                            order by cod_role_cuenta, desc_cuenta""" % (cod_documento, cod_role))
    data = dictfetchall(cursor)
    cursor.execute("""select desc_role, cod_role_cuenta,desc_escenario, desc_dimension, desc_tipo_cuenta,coalesce(desc_role_cuenta, etiqueta, desc_cuenta) desc_cuenta, cod_periodo,desc_periodo, desc_detalle_documento, decimales, cod_unidad, desc_unidad, cod_entidad,desc_entidad, case when not decimales isnull then desc_detalle_documento::numeric / 10 ^ abs(decimales) else 0 end valor 
                                from detalle_documento 
                                join cuenta using (cod_cuenta) 
                                left join role_cuenta using (cod_cuenta)
                                left join role using(cod_role)
                                left join unidad using (cod_unidad)
                                join tipo_cuenta using (cod_tipo_cuenta)
                                join contexto using (cod_contexto) 
                                join entidad using (cod_entidad)
                                join periodo using (cod_periodo) 
                                left join context_escenarios using (cod_contexto) 
                                left join escenario using (cod_escenario) 
                                left join dimension using (cod_dimension) 
                                --where cod_periodo = 44 and cod_tipo_cuenta in (6,7)
                                where cod_dimension is null  and cod_documento = %s and cod_role =%s
                                order by cod_role_cuenta, desc_cuenta""" % (cod_documento, cod_role))
    date = cursor.fetchall()
    cursor.execute("""select distinct desc_periodo 
                               from detalle_documento 
                               join cuenta using (cod_cuenta) 
                               left join role_cuenta using (cod_cuenta)
                               left join role using(cod_role)
                               join contexto using (cod_contexto) 
                               join periodo using (cod_periodo)
                               left join context_escenarios using (cod_contexto) 
                               left join escenario using (cod_escenario) 
                               left join dimension using (cod_dimension)
                               where cod_dimension is null  and cod_documento = %s and cod_role =%s
                               order by desc_periodo desc""" % (cod_documento, cod_role))
    per = dictfetchall(cursor)
    for d in date:
        cod_rol = d[1]
        desc_role = d[0]
        entidad = d[13]
        cuenta = d[5]
        dineros = d[8]
        fechas = d[7]
    hoy = datetime.now()
    hoy = hoy.strftime("%Y-%m-%d")
    report = PdfCustomDetail(filename="%s.pdf" % (cod_rol), title=["%s" % (desc_role), entidad.__str__(), cuenta],
                             logo=getattr(sys, 'logo', 'xeeffy_logo_index.png'))
    pvt_kms = pivot(data, ('cod_role_cuenta', 'desc_cuenta'), ('desc_periodo',), 'valor')
    datas = []
    columnas = ["CUENTA"]
    data_linea = ['']
    ancho_total = 0
    colsWidth = [0]
    colAlign = ['LEFT']
    colType = ['str']
    for p in per:
        columnas.append(p['desc_periodo'])
        data_linea.append(0.0)
        colsWidth.append(60)
        ancho_total += 60
        colAlign.append("RIGHT")
        colType.append("str")
    colsWidth[0] = 540 - ancho_total
    for f in pvt_kms:
        dl = list(f.get(('cod_role_cuenta', 'desc_cuenta'), ['']))
        for p in per:
            dl.append(f.get((p['desc_periodo'],), 0))
        datas.append(dl)
    data_final = []
    datas.sort()
    for d in datas:
        data_final.append(d[1:])
    return render_template("/balance.html", d=data_final, columnas=columnas, entidad=entidad, desc_role=desc_role,
                           hoy=hoy)


# funciones de ordenado y calculos de cada uno de los servicios
def JsonAgg(json, valores, claves, funcs=None):
    aggby = claves
    vals = valores
    js = json
    result = {}
    jresult = []
    nVals = len(valores)
    if funcs == None:
        funcs = len(valores) * ["sum"]
    for d in js:
        k = tuple([(i, d[i]) for i in aggby])
        result[k] = result.get(k, {})
        for idxv in range(nVals):
            va = vals[idxv]
            result[k][va] = result[k].get(va, 0)
            if funcs[idxv] == 'sum':
                result[k][va] += d.get(va)
            elif funcs[idxv] == 'count':
                result[k][va] += 1
            elif funcs[idxv] == 'max':
                if d.get(va) > result[k][va]:
                    result[k][va] = d.get(va)
            elif funcs[idxv] == 'min':
                if d.get(va) < result[k][va]:
                    result[k][va] = d.get(va)
    for r in result:
        jresult.append(dict(dict(r), **result[r]))
    del result
    return jresult


def CRut(rut):
    if rut == "":
        return rut
    rut = str.replace(rut, ".", "")
    rut = str.replace(rut, "-", "")
    rut = "0000000000" + rut
    l = len(rut)
    rut_aux = "-" + rut[l - 1:l]
    l = l - 1
    while 2 < l:
        rut_aux = "." + rut[l - 3:l] + rut_aux
        l = l - 3

    rut_aux = rut[0:l] + rut_aux
    l = len(rut_aux)
    rut_aux = rut_aux[l - 12:l]
    return rut_aux


if __name__ == "__main__":
    db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5555)
